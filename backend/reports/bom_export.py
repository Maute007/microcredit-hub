"""Generate BdM Ficha de Reporte Trimestral as Excel."""
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generate_bom_report(settings, loans_qs, payments_qs, clients_qs, date_from, date_to, period_label):
    """
    Generate the BdM report Excel workbook.
    Returns an in-memory bytes object.

    Arguments:
    - settings: SystemSettings instance
    - loans_qs: Loan queryset (all loans)
    - payments_qs: Payment queryset (all payments)
    - clients_qs: Client queryset (all clients)
    - date_from, date_to: datetime.date objects for the reporting period
    - period_label: string like "Janeiro de 2026"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "FICHA DE REPORTE TRIMESTRAL"

    # ---- Set column widths ----
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # ---- Styles ----
    gray_fill = PatternFill("solid", fgColor="4472C4")
    light_fill = PatternFill("solid", fgColor="DCE6F1")

    row = 1

    def write_row(label, c=None, d=None, e=None, bold=False, fill=None):
        nonlocal row
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name="Arial", bold=bold, size=9)
        if fill:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
        if c is not None:
            ws.cell(row=row, column=3, value=c).font = Font(name="Arial", size=9)
        if d is not None:
            ws.cell(row=row, column=4, value=d).font = Font(name="Arial", size=9)
        if e is not None:
            ws.cell(row=row, column=5, value=e).font = Font(name="Arial", size=9)
        row += 1

    # ---- HEADER ----
    ws.cell(row=row, column=1, value="BANCO DE MOÇAMBIQUE").font = Font(name="Arial", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="DEPARTAMENTO DE SUPERVISÃO PRUDENCIAL").font = Font(name="Arial", bold=True, size=10)
    row += 1
    ws.cell(row=row, column=1, value="REPORTE PERIÓDICO DE INFORMAÇÕES DE MICROFINANÇAS").font = Font(name="Arial", bold=True, size=10)
    row += 1
    ws.cell(row=row, column=1, value="INSTITUIÇÕES SUJEITAS À MONITORIZAÇÃO").font = Font(name="Arial", bold=True, size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"PERÍODO DE REPORTE: {period_label}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"DATA: {date.today().strftime('%d/%m/%Y')}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=4, value="(Valores em Meticais)").font = Font(name="Arial", italic=True, size=8)
    row += 1

    # ---- INSTITUTION INFO ----
    ws.cell(row=row, column=1, value="OPERADORES DE MICROCRÉDITO").font = Font(name="Arial", bold=True, size=9)
    row += 1
    institution_name = getattr(settings, "creditor_legal_name", "") or getattr(settings, "name", "") or ""
    ws.cell(row=row, column=1, value=f"Denominação: {institution_name}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"Endereço: {getattr(settings, 'creditor_address', '') or ''}").font = Font(name="Arial", size=9)
    row += 1
    province = getattr(settings, "bom_province", "") or getattr(settings, "creditor_city", "") or ""
    ws.cell(row=row, column=1, value=f"Provincia: {province}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"Telefone: {getattr(settings, 'bom_phone', '') or ''}").font = Font(name="Arial", size=9)
    row += 1
    fax = getattr(settings, "bom_fax", "") or "________________"
    bom_email = getattr(settings, "bom_email", "") or ""
    ws.cell(row=row, column=1, value=f"Fax: {fax}     E-mail: {bom_email}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"Nº de Trabalhadores: {getattr(settings, 'bom_num_workers', 1) or 1}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"Data de Início das Actividades: {getattr(settings, 'bom_start_date', '') or ''}").font = Font(name="Arial", size=9)
    row += 1
    ws.cell(row=row, column=1, value=f"Nome do Operador: {getattr(settings, 'bom_operator_name', '') or ''}").font = Font(name="Arial", size=9)
    row += 1
    row += 1

    # ---- Section header helpers ----
    def section_header(text):
        nonlocal row
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = gray_fill
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        row += 1

    def sub_header(text, c_label=None, d_label=None, e_label=None):
        nonlocal row
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = light_fill
        ws.cell(row=row, column=1, value=text).font = Font(name="Arial", bold=True, size=9)
        if c_label:
            ws.cell(row=row, column=3, value=c_label).font = Font(name="Arial", bold=True, size=9)
        if d_label:
            ws.cell(row=row, column=4, value=d_label).font = Font(name="Arial", bold=True, size=9)
        if e_label:
            ws.cell(row=row, column=5, value=e_label).font = Font(name="Arial", bold=True, size=9)
        row += 1

    # ---- Calculate data ----
    from django.db.models import Sum, Min, Max

    # 2.1.1 — Volume de créditos
    # Loans created in period (using start_date as proxy for creation in period)
    loans_period = loans_qs.filter(start_date__gte=date_from, start_date__lte=date_to)
    concedidos_capital = loans_period.aggregate(s=Sum("amount"))["s"] or Decimal("0")
    concedidos_total = loans_period.aggregate(s=Sum("total_amount"))["s"] or Decimal("0")
    concedidos_juro = concedidos_total - concedidos_capital

    # Reembolsados = paid payments in period
    paid_in_period = payments_qs.filter(
        status="pago",
        date__gte=date_from,
        date__lte=date_to,
    )
    reembolsados_capital = paid_in_period.aggregate(s=Sum("amount"))["s"] or Decimal("0")

    # Abatidos = loans fully paid in period (status pago, end_date within period as proxy)
    abatidos_loans = loans_qs.filter(status="pago", end_date__gte=date_from, end_date__lte=date_to)
    abatidos_capital = abatidos_loans.aggregate(s=Sum("amount"))["s"] or Decimal("0")
    abatidos_total = abatidos_loans.aggregate(s=Sum("total_amount"))["s"] or Decimal("0")

    # Active portfolio
    active_loans = loans_qs.filter(status="ativo")
    carteira_capital = active_loans.aggregate(s=Sum("amount"))["s"] or Decimal("0")
    carteira_total = active_loans.aggregate(s=Sum("total_amount"))["s"] or Decimal("0")

    # Portfolio at risk = overdue loans
    risk_loans = loans_qs.filter(status="atrasado")
    risco_capital = risk_loans.aggregate(s=Sum("amount"))["s"] or Decimal("0")

    # 2.1.2 — Number of loans
    num_concedidos = loans_period.count()
    num_reembolsados = loans_qs.filter(status="pago", end_date__gte=date_from, end_date__lte=date_to).count()

    # 2.1.3 — By sector
    SECTORS = ["comercio", "agricultura", "pecuaria", "industria", "servicos", "consumo", "outros"]
    sector_data = {}
    for s in SECTORS:
        agg = loans_period.filter(sector=s).aggregate(s=Sum("amount"))
        sector_data[s] = agg["s"] or Decimal("0")
    total_sector = sum(sector_data.values())
    total_all_capital = loans_qs.aggregate(s=Sum("amount"))["s"] or Decimal("0")

    # 2.1.4 — Client portfolio by gender
    active_clients = clients_qs.filter(status="ativo")
    men = active_clients.filter(gender="M").count()
    women = active_clients.filter(gender="F").count()
    others_gender = active_clients.filter(gender="O").count()
    total_clients = active_clients.count()

    # 2.1.5 — Risk structure (overdue payments by days past due)
    today = date.today()
    class1 = {"capital": Decimal("0"), "juro": Decimal("0"), "total": Decimal("0")}
    class2 = {"capital": Decimal("0"), "juro": Decimal("0"), "total": Decimal("0")}
    class3 = {"capital": Decimal("0"), "juro": Decimal("0"), "total": Decimal("0")}
    class4 = {"capital": Decimal("0"), "juro": Decimal("0"), "total": Decimal("0")}

    # Use overdue payments to classify by how overdue they are
    overdue_payments = payments_qs.filter(status="atrasado").select_related("loan")
    for pmt in overdue_payments:
        # Use the payment date as the reference for days overdue
        if not pmt.date:
            continue
        days = (today - pmt.date).days
        cap = pmt.amount or Decimal("0")
        if 1 <= days <= 30:
            class1["capital"] += cap
        elif 31 <= days <= 90:
            class2["capital"] += cap
        elif 91 <= days <= 365:
            class3["capital"] += cap
        elif days > 365:
            class4["capital"] += cap

    for c in [class1, class2, class3, class4]:
        c["total"] = c["capital"] + c["juro"]

    # 2.2 — Interest rates and terms
    active_rates = loans_qs.filter(status__in=["ativo", "atrasado"]).aggregate(
        min_rate=Min("interest_rate"),
        max_rate=Max("interest_rate"),
        min_term=Min("term"),
        max_term=Max("term"),
    )

    # ---- Write sections ----
    section_header("2. INFORMAÇÕES SOBRE A ACTIVIDADE")
    section_header("2.1 CARTEIRA DE CRÉDITO")
    sub_header("2.1.1 Volume de créditos - MZN", "Capital", "Juro", "TOTAL")

    write_row("Montante de créditos concedidos no período", float(concedidos_capital), float(concedidos_juro), float(concedidos_total))
    write_row("Montante de créditos reembolsados no período", float(reembolsados_capital), 0, float(reembolsados_capital))
    write_row("Montante de créditos abatidos no período", float(abatidos_capital), 0, float(abatidos_total))
    write_row("Montante da carteira de crédito activa/vigente:", float(carteira_capital), 0, float(carteira_total))
    write_row("Montante da carteira em risco", float(risco_capital), None, float(risco_capital))
    row += 1

    sub_header("2.1.2 Número de créditos")
    write_row("Número de créditos concedidos no período", None, None, num_concedidos)
    write_row("Número de créditos reembolsados no período", None, None, num_reembolsados)
    row += 1

    sub_header("2.1.3 Créditos concedidos por sector de actividade/finalidade", None, None, "Montante de créditos concedidos no período")
    sector_labels = {
        "comercio": "Comércio",
        "agricultura": "Agricultura",
        "pecuaria": "Pecuária",
        "industria": "Indústria",
        "servicos": "Serviços",
        "consumo": "Consumo",
        "outros": "Outros",
    }
    for s_key, s_label in sector_labels.items():
        write_row(s_label, None, None, float(sector_data.get(s_key, 0)))
    write_row("TOTAL", float(total_all_capital), None, float(total_sector), bold=True)
    row += 1

    sub_header("2.1.4 Carteira de Clientes", None, None, "Nº de clientes activos")
    write_row("Homens", None, None, men)
    write_row("Mulheres", None, None, women)
    write_row("Outros", None, None, others_gender)
    write_row("Total", None, None, total_clients, bold=True)
    row += 1

    sub_header("2.1.5 Estrutura da Carteira em Risco", "Capital", "Juro", "TOTAL")
    write_row("Classe I (de 1 até 30 dias)", float(class1["capital"]), float(class1["juro"]), float(class1["total"]))
    write_row("Classe II (de 31 até 90 dias)", float(class2["capital"]), float(class2["juro"]), float(class2["total"]))
    write_row("Classe III (de 91 a 365 dias)", float(class3["capital"]), float(class3["juro"]), float(class3["total"]))
    write_row("Classe IV (mais de 365 dias)", float(class4["capital"]), float(class4["juro"]), float(class4["total"]))
    total_risk_cap = class1["capital"] + class2["capital"] + class3["capital"] + class4["capital"]
    total_risk_juro = class1["juro"] + class2["juro"] + class3["juro"] + class4["juro"]
    write_row("Total", float(total_risk_cap), float(total_risk_juro), float(total_risk_cap + total_risk_juro), bold=True)
    row += 1

    section_header("2.2 TAXAS DE JURO E PRAZOS DE VENCIMENTO")
    sub_header("Descrição", None, "Mínimo", "Máximo")
    min_rate = active_rates["min_rate"] or Decimal("0")
    max_rate = active_rates["max_rate"] or Decimal("0")
    min_term = active_rates["min_term"] or 1
    max_term = active_rates["max_term"] or 1
    write_row("Taxa de juro mensal", None, f"{min_rate:.2f}%", f"{max_rate:.2f}%")
    write_row("Prazo (Meses)", None, f"{min_term} Mês", f"{max_term} Meses")
    row += 1

    own_capital = getattr(settings, "bom_own_capital", Decimal("0")) or Decimal("0")
    foreign_national = getattr(settings, "bom_foreign_capital_national", Decimal("0")) or Decimal("0")
    foreign_foreign = getattr(settings, "bom_foreign_capital_foreign", Decimal("0")) or Decimal("0")
    financing_loans = getattr(settings, "bom_financing_loans", Decimal("0")) or Decimal("0")
    financing_donations = getattr(settings, "bom_financing_donations", Decimal("0")) or Decimal("0")
    financing_capital_increase = getattr(settings, "bom_financing_capital_increase", Decimal("0")) or Decimal("0")
    initial_capital = getattr(settings, "bom_initial_capital", Decimal("0")) or Decimal("0")
    current_capital = getattr(settings, "bom_current_capital", Decimal("0")) or Decimal("0")

    section_header("2.3 FONTES DE FINANCIAMENTO DA INSTITUIÇÃO")
    write_row("Capitais Próprios", float(own_capital))
    write_row("Capitais Alheios")
    write_row("    Nacionais", float(foreign_national))
    write_row("    Estrangeiros", float(foreign_foreign))
    total_fin = own_capital + foreign_national + foreign_foreign
    write_row("Total", float(total_fin), bold=True)
    row += 1

    section_header("2.4 FINANCIAMENTOS DO PERÍODO")
    write_row("Empréstimos obtidos no período", float(financing_loans))
    write_row("Donativos obtidos no período", float(financing_donations))
    write_row("Aumento do capital com recursos próprios", float(financing_capital_increase))
    total_f = financing_loans + financing_donations + financing_capital_increase
    write_row("Total", float(total_f), bold=True)
    row += 1

    section_header("2.5 CAPITAL")
    write_row("Inicial", float(initial_capital))
    write_row("Actual", float(current_capital))
    row += 1

    section_header("3. SITUAÇÃO FINANCEIRA DA OPERADOR")
    sub_header("", "Mês 1", "Mês 2", "Mês 3")
    fin_sit = list(getattr(settings, "bom_financial_situation", None) or [])
    while len(fin_sit) < 3:
        fin_sit.append({"caixa": 0, "bancos": 0, "outros_activos": 0})

    write_row(
        "Caixa",
        float(fin_sit[0].get("caixa", 0)),
        float(fin_sit[1].get("caixa", 0)),
        float(fin_sit[2].get("caixa", 0)),
    )
    write_row(
        "Bancos",
        float(fin_sit[0].get("bancos", 0)),
        float(fin_sit[1].get("bancos", 0)),
        float(fin_sit[2].get("bancos", 0)),
    )
    write_row(
        "Outros Activos",
        float(fin_sit[0].get("outros_activos", 0)),
        float(fin_sit[1].get("outros_activos", 0)),
        float(fin_sit[2].get("outros_activos", 0)),
    )
    totals_fin = [
        fin_sit[i].get("caixa", 0) + fin_sit[i].get("bancos", 0) + fin_sit[i].get("outros_activos", 0)
        for i in range(3)
    ]
    write_row("Total", float(totals_fin[0]), float(totals_fin[1]), float(totals_fin[2]), bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
